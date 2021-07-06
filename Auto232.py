import openpyxl  # Needed for Excel
import re  # Needed for special splits and sub
import time  # Needed for dates
from datetime import timedelta  # Needed for time calculations
# Approved
def readFile(logs):
    """Extract the log data and make it usable."""
    """
    logs - list of list of str, empty, will contain all session information
    """
    with open("BeltMovement", "r") as file:
        # Use 'with open' for automatic file closing
        Lines = file.readlines()
        temp = ""  # Meant for the first fragmented entry
        for line in Lines:
            # Any [0:-1] in this function is for removing the newline chr
            if (line == "\n") or (line.__contains__("*")): pass
                # Skip empty lines and * placeholder lines
            elif line.__contains__("Termite log"): logs.append([line[0:-1]])
                # Add a new session when the log marks a new session
            elif ':' not in line: temp = line[0:-1]
                # Store the fragmented entry for later
            elif temp:
                # Gather the fragments and record the complete entry
                logs[-1].append(line.split(" ", 1)[0] + " " + temp)
                # Record the currently read in complete entry
                logs[-1].append(line.split(" ", 1)[1][0:-1])
                temp = ""  # Clear the stored fragment
            else: logs[-1].append(line[0:-1])
                # Record the entry

    for log in logs:  # Check each session
        checkValidity(log)
# Approved
def checkValidity(log):
    """Make sure there's nothing syntactically wrong with the data."""
    """
    log - list of str, contains session info and entries
    """
    state = ""  # Used to make sure the 1's and 0's alternate
    pTime = ""  # Holds previous time
    cTime = ""  # Holds current time

    if not log[0].__contains__("Termite log"):
        # Execute if the session doesn't start with this
        raise Exception("Error in a session info stamp")
    if len(log) == 1: return  # A stamp but no entries

    for line in log[1:]:  # Starting after the session info stamp
        entry = line.split()  # Separate time stamp, data, and time
        if not state:  # Set state depending on first data entry
            if entry[1] == "1": state = "Low"
            elif entry[1] == "0": state = "High"
        # Check time stamps
        if line == log[1]:  # Set pTime on first entry
                pTime = entry[0]
        else:  # Update cTime
            cTime = entry[0]
            if pTime > cTime:  # If pTime is later than cTime
                raise Exception("Time error in session: {}".format(log[0]))
            else: pTime = cTime
                # Update pTime when check passed
        # Check data
        if not state: continue  # Waiting for a state
        if (entry[1] == "1") and (state == "Low"): state = "High"
            # Confirm a 1, now expecting a 0 next
        elif (entry[1] == "0") and (state == "High"): state = "Low"
            # Confirm a 0, now expecting a 1 next
        elif entry[1] == "3": continue  # To prevent indexing error
        else:  # Occurs if the transmitted data is unrecognized
            raise Exception("Unrecognized data in session: {}".format(log[0]))
        # Check time
        if (len(entry[2]) != 6  # 6 doesn't include newline
            or entry[2][2] != '.'
            or not entry[2][0:2].isdigit()
            or not entry[2][3:].isdigit()):
            raise Exception("Incorrect time mode format in session: {}".format(log[0]))
# Approved
def analyzeData(log):
    """Return when/for how long the belt moved and stopped."""
    """
    log - list of str, contains session info and entries
    """
    # Approved
    def moveDetector(diff, period, moveBlocks, start, pTime):
        """Check if the belt is moving or stopped."""
        """
        diff - float, amount of time passed since last data
        period - diff, amount seconds to compare diff to 
        moveBlocks - list of list of str, stores intervals of movement
        start - str, time when the belt started moving
        pTime - str, previous time
        """
        if float(diff) < period:  # If smoovin
            if not start:  # Start of movement interval
                start = pTime  # Set start
                moveBlocks.append([start])  # Record start of interval
        else:  # If stoppin
            if start:  # End of movement interval
                moveBlocks[-1].append(pTime)  # Record end of interval
                start = ""  # Clear start
        return start

    begin = ""  # Holds time that starts a state
    end = ""  # Holds time that ends a state, doubles as current time
    prevTime = ""  # Holds previous time
    SmooveBlocks = []
    for line in log[1:]:  # Go through each entry
        entry = line.split()  # Separate time and data
        if line == log[1]: prevTime = entry[0][0:-1]  # Exclude ending ':'
            # On the first entry, set prevTime
        else:  # Update end
            end = entry[0][0:-1]
            if entry[1] == "1":  # Reached end of chain/start of gap
                begin = moveDetector(entry[2], 5.3125, SmooveBlocks, begin, prevTime)
            elif entry[1] == "0":  # Reached end of gap/start of chain
                begin = moveDetector(entry[2], 4.25, SmooveBlocks, begin, prevTime)
            elif entry[1] == "3":  # Instant stoppage
                begin = moveDetector(1, 0, SmooveBlocks, begin, prevTime)
            else: pass
                # Stoppage
            prevTime = end  # Update prevTime

        if line == log[-1] and len(SmooveBlocks[-1]) == 1:
            # On last entry and half interval
            SmooveBlocks[-1].append(end)

    return SmooveBlocks
# Approved
def toExcel(logs, xlData):
    """Make Excel workbook with data."""
    """
    logs - list of list of str, contains all session information
    xlData - list of list of str, moving intervals for every session
    """
    def timeDiff(aTime, bTime):
        """Return the difference in times."""
        """
        aTime - str, an earlier time
        bTime - str, a later time
        """
        aTemp = aTime.split(":")
        bTemp = bTime.split(":")
        aTD = timedelta(hours=int(aTemp[0]), minutes=int(aTemp[1]), seconds=float(aTemp[2]))
        bTD = timedelta(hours=int(bTemp[0]), minutes=int(bTemp[1]), seconds=float(bTemp[2]))
        c = bTD - aTD
        min = c.total_seconds() // 60  # Total minutes without remainder
        r_sec = c.total_seconds() % 60  # Remaining seconds
        hr = min // 60  # Total hours without remainder
        r_min = min % 60  # Remaining minutes
        return [hr, r_min, r_sec]
    def tolTime(day):
        """Return the daily total of movement time."""
        """
        day - list of list of float, each list is an interval's duration
        """
        total_hr = 0
        total_min = 0
        total_s = 0.0
        for session in day:
            total_hr += session[0]
            total_min += session[1]
            total_s += session[2]
        total_min += total_s // 60  # Take any minutes from seconds
        total_s = total_s % 60  # Adjust seconds
        total_hr += total_min // 60  # Take any hours from minutes
        total_min = total_min % 60  # Adjust minutes

        return "Daily Total: {}:{}:{}".format(total_hr, total_min, total_s)

    wb = openpyxl.Workbook()  # Create a new workbook
    ws = wb.active  # Set only sheet as active
    row_num = 1  # Will change to cover many rows, starts as first row
    ws.cell(row=row_num, column=1).value = "All intervals represent periods of movement"
    ws.cell(row=row_num+1, column=1).value = "HH:MM:SS.SS in military time"
    row_num += 3  # Get ready to write in sessions
    for x in range(len(logs)):  # For each session
        stamp = re.split("Termite log, started at ", logs[x][0])[1]  # Time stamp
        ws.cell(row=row_num, column=1).value = stamp
        row_num += 1  # Get ready to write in data
        times = []  # Needed for daily total
        data = xlData[x]  # Data for only one session
        if data:  # If there's something
            count = len(data)  # Amount of intervals
            for i in range(count):  # For each interval
                B = data[i][0]  # Beginning of interval
                E = data[i][1]  # Ending of interval
                ws.cell(row=row_num+i, column=1).value = B + "-" + E
                inter = timeDiff(B, E)  # Get duration
                times.append(inter)
                ws.cell(row=row_num+i, column=4).value = ("Hours: {}, Minutes: {}, Seconds: {}"
                                                          .format(inter[0], inter[1], inter[2]))
            ws.cell(row=row_num+count, column=4).value = tolTime(times)  # Daily total
            row_num += count + 1  # Update row_num to pass all recently filled rows
        else: row_num += 1
    
    date = time.strftime("%D", time.localtime())  # Get MM/DD/YYYY
    date = re.sub("/", "_", date)  # Replace / with _ for valid name
    name = "AutoLine_Weekly_Report_" + date + ".xlsx"
    wb.save(name)  # Save changes
# Approved
def main():
    logs = []
    xlData = []
    readFile(logs)
    for log in logs:  # Cycle through each session
        xlData.append(analyzeData(log))
    toExcel(logs, xlData)

main()
